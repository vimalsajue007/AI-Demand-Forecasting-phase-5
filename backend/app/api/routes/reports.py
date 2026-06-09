import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.db.database import get_db
from app.models.user import User
from app.models.forecast import Forecast
from app.core.security import get_current_user
from app.core.roles import require_permission
from app.services.notification_service import notify_report_generated
from app.services.activity_service import log_activity

router = APIRouter(prefix="/api/reports", tags=["Reports"])


@router.get("/{forecast_id}/excel")
def export_excel(forecast_id: int, db: Session = Depends(get_db),
                 current_user: User = Depends(require_permission("download_reports"))):
    forecast = db.query(Forecast).filter(Forecast.id == forecast_id, Forecast.owner_id == current_user.id).first()
    if not forecast:
        raise HTTPException(status_code=404, detail="Forecast not found")
    if forecast.status != "completed":
        raise HTTPException(status_code=400, detail="Forecast not completed yet")

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    green_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")

    ws["A1"] = "Forecast Report"
    ws["A1"].font = Font(bold=True, size=14)
    for row_idx, (label, value) in enumerate([
        ["Forecast Name", forecast.name], ["Model", forecast.model_type],
        ["Periods", forecast.periods], ["Accuracy", f"{round(forecast.accuracy_score * 100, 2)}%" if forecast.accuracy_score else "N/A"],
        ["MAE", round(forecast.mae, 4) if forecast.mae else "N/A"],
        ["RMSE", round(forecast.rmse, 4) if forecast.rmse else "N/A"],
        ["Status", forecast.status], ["Created", str(forecast.created_at)[:19]],
    ], start=3):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)

    if forecast.historical_data:
        ws_hist = wb.create_sheet("Historical")
        headers = list(forecast.historical_data[0].keys())
        for ci, h in enumerate(headers, 1):
            cell = ws_hist.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = green_fill
        for ri, row in enumerate(forecast.historical_data, 2):
            for ci, h in enumerate(headers, 1):
                ws_hist.cell(row=ri, column=ci, value=row.get(h))

    if forecast.predictions:
        ws_pred = wb.create_sheet("Predictions")
        headers = list(forecast.predictions[0].keys())
        for ci, h in enumerate(headers, 1):
            cell = ws_pred.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = green_fill
        for ri, row in enumerate(forecast.predictions, 2):
            for ci, h in enumerate(headers, 1):
                ws_pred.cell(row=ri, column=ci, value=row.get(h))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    notify_report_generated(db, current_user.id, forecast.name, "excel")
    log_activity(db, "report_downloaded_excel", user_id=current_user.id, resource="forecast", resource_id=forecast_id)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="forecast_{forecast_id}.xlsx"'})


@router.get("/{forecast_id}/pdf")
def export_pdf(forecast_id: int, db: Session = Depends(get_db),
               current_user: User = Depends(require_permission("download_reports"))):
    forecast = db.query(Forecast).filter(Forecast.id == forecast_id, Forecast.owner_id == current_user.id).first()
    if not forecast:
        raise HTTPException(status_code=404, detail="Forecast not found")
    if forecast.status != "completed":
        raise HTTPException(status_code=400, detail="Forecast not completed")

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    green = HexColor("#166534")
    light_green = HexColor("#DCFCE7")

    elements = []
    title_style = ParagraphStyle("title", parent=styles["Title"], textColor=green, fontSize=18)
    elements.append(Paragraph("AI Demand Forecasting Report", title_style))
    elements.append(Spacer(1, 0.2 * inch))

    info_data = [
        ["Forecast Name:", forecast.name], ["Model:", forecast.model_type.replace("_", " ").title()],
        ["Periods:", str(forecast.periods)],
        ["Accuracy (R²):", f"{round(forecast.accuracy_score * 100, 2)}%" if forecast.accuracy_score else "N/A"],
        ["MAE:", str(round(forecast.mae, 4)) if forecast.mae else "N/A"],
        ["RMSE:", str(round(forecast.rmse, 4)) if forecast.rmse else "N/A"],
        ["Status:", forecast.status.upper()], ["Generated:", str(forecast.created_at)[:19]],
    ]
    table = Table(info_data, colWidths=[2 * inch, 4 * inch])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (0,-1), light_green), ("TEXTCOLOR", (0,0), (0,-1), green),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"), ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("PADDING", (0,0), (-1,-1), 6),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.3 * inch))

    if forecast.predictions:
        elements.append(Paragraph("Forecast Predictions", styles["Heading2"]))
        pred_data = forecast.predictions[:20]
        headers = list(pred_data[0].keys())
        table_data = [headers] + [[str(row.get(h, "")) for h in headers] for row in pred_data]
        pred_table = Table(table_data)
        pred_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), green), ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ("PADDING", (0,0), (-1,-1), 4), ("FONTSIZE", (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, light_green]),
        ]))
        elements.append(pred_table)

    doc.build(elements)
    buffer.seek(0)
    notify_report_generated(db, current_user.id, forecast.name, "pdf")
    log_activity(db, "report_downloaded_pdf", user_id=current_user.id, resource="forecast", resource_id=forecast_id)
    return StreamingResponse(buffer, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="forecast_{forecast_id}.pdf"'})


@router.get("/{forecast_id}/insights")
def get_ai_insights(forecast_id: int, db: Session = Depends(get_db),
                    current_user: User = Depends(require_permission("view_reports"))):
    forecast = db.query(Forecast).filter(Forecast.id == forecast_id, Forecast.owner_id == current_user.id).first()
    if not forecast:
        raise HTTPException(status_code=404, detail="Forecast not found")
    if forecast.status != "completed":
        raise HTTPException(status_code=400, detail="Forecast not completed")

    insights = []
    accuracy = forecast.accuracy_score * 100 if forecast.accuracy_score else 0

    if accuracy >= 80:
        insights.append({"type":"success","title":"High Model Accuracy","message":f"{forecast.model_type.replace('_',' ')} achieved {accuracy:.1f}% accuracy — excellent for business planning."})
    elif accuracy >= 60:
        insights.append({"type":"warning","title":"Moderate Accuracy","message":f"Accuracy is {accuracy:.1f}%. Consider trying the Ensemble model."})
    else:
        insights.append({"type":"error","title":"Low Accuracy","message":f"Accuracy is {accuracy:.1f}%. Collect more data or use the Ensemble model."})

    if forecast.predictions:
        import numpy as np
        preds = [p["yhat"] for p in forecast.predictions]
        trend = "upward" if preds[-1] > preds[0] else "downward"
        pct = abs((preds[-1] - preds[0]) / (preds[0] + 1e-8)) * 100
        insights.append({"type":"info","title":f"Demand Trend: {trend.title()}","message":f"Demand is forecasted to {trend} by {pct:.1f}% over {forecast.periods} periods."})
        if forecast.mae:
            insights.append({"type":"info","title":"Forecast Precision","message":f"Average error (MAE) is {forecast.mae:.2f} units."})

    return {"forecast_id": forecast_id, "forecast_name": forecast.name, "insights": insights,
            "summary": f"Forecast '{forecast.name}' — {accuracy:.1f}% accuracy using {forecast.model_type.replace('_',' ')} model."}
